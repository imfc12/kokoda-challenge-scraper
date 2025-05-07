import kokoda_scraper as ks
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.utils.cell import rows_from_range, range_boundaries, get_column_letter
from datetime import datetime
from dotenv import load_dotenv
import os
import ast

# Determine workbook path that holds 'Kokoda Results.xlsx'
wb_path = 'example_output/Kokoda Results.xlsx'

load_dotenv()
# Retrieve team names from .env, private info
lords_team_names = os.getenv('team_names_env').split(',')

# Team ranges also includes team names, private info
lords_teams_ranges_str = os.getenv('team_ranges_env')
lords_teams_ranges = ast.literal_eval(lords_teams_ranges_str)

# Create the worksheet for today's date
# Pass in a boolean that toggles TEST sheet, if True, creates a 'TEST' sheet
def define_excel_objects(date_bool) -> tuple[Workbook, Worksheet]:
    operating_wb = load_workbook(wb_path)
    # Load the current worksheets
    current_sheets = operating_wb.sheetnames
    if date_bool:
        if 'TEST' in current_sheets:
            operating_ws = operating_wb['TEST']
            return operating_wb, operating_ws
        source_ws = operating_wb['Template']
        operating_ws = operating_wb.copy_worksheet(source_ws)
        operating_ws.title = 'TEST'
        return operating_wb, operating_ws
    
    time = datetime.now()
    # Create the sheet name for today's date
    current_day_sheet_name = f'{time.day}-{time.month}'
    
    # If the current sheet name exists i.e. script and output has already been done once today...
    if current_day_sheet_name in current_sheets:
        # We override the current day's sheet
        operating_ws = operating_wb[current_day_sheet_name]
    else:
        # Else we create a new sheet by copying the template and naming the sheet after today's date
        source_ws = operating_wb['Template']
        operating_ws = operating_wb.copy_worksheet(source_ws)
        operating_ws.title = current_day_sheet_name

    return operating_wb, operating_ws

def excel_transfer(teams: list[str], date_bool: bool = False) -> None:
    # Retrieve allocated workbook and worksheet for the Kokoda data to go into
    wb, ws = define_excel_objects(date_bool)
    # Return the team's details dictionary, overall_dictionary (top team & member fundraiser), and overall_fundraising variable (integer)
    teams_dict, overall_dict, overall_fundraising = ks.process_kokoda_data(teams)
    for team_name, info in teams_dict.items():
        # print(team_name)
        # Info is a dictionary, returns 5 key:value pairs as shown below
        raised = info['rais']
        target = info['targ']
        percentage = info['perc']
        members = info['memb']
        rank = info['rank'] # Currently unused variable
        # Retrieve row pair for the given range eg. range = 'A1:B2' -> returns (A1, B1), (A2, B2) etc. creates an iterator. Yields one row at a time
        team_cell_row_range = rows_from_range(lords_teams_ranges[team_name])
        # Iterate over members list and range rows at the same time to plot in corresponding cells
        # 'members' is list of tuples eg. [('Mike', 300), ('Craig', 120)]
        for mem, row in zip(members, team_cell_row_range):
            ws[row[0]] = mem[0]
            ws[row[1]] = mem[1]

        # Range boundaries returns index of (min_col, min_row, max_col, max_row). Retrieve last column and last row
        max_col_value = range_boundaries(lords_teams_ranges[team_name])[2]
        max_row_value = range_boundaries(lords_teams_ranges[team_name])[3]
        # Retrieve member money column
        mm_col = get_column_letter(max_col_value)
        # Add numbers to max row to iterate over cells
        ws[f'{mm_col}{max_row_value + 1}'] = raised
        ws[f'{mm_col}{max_row_value + 2}'] = target
        # Convert percentage into usable value for excel
        ws[f'{mm_col}{max_row_value + 3}'] = percentage / 100

    # Place top fundraising team
    ws['I2'] = f'{overall_dict['top_team_fundraiser'][0]} - ${overall_dict['top_team_fundraiser'][1]}'
    # Place top fundraising member
    ws['I3'] = f'{overall_dict['top_mem_fundraiser'][0]} - ${overall_dict['top_mem_fundraiser'][1]} - {overall_dict['top_mem_fundraiser'][2]}'
    # Plot overall fundraising into cell (all teams fundraising combined)
    ws['C3'] = overall_fundraising
    
    # Insert time of execution into sheet
    time = datetime.now()
    ws['E2'] = time.date()
    ws['E3'] = time.time()

    wb.save(wb_path)
    print('Complete')   


if __name__ == '__main__':
    # True = 'TEST' sheet
    # If 'True' is passed in we want to create a test sheet for results
    excel_transfer(teams=lords_team_names, date_bool=False)





